import cv2
import sqlite3
import numpy as np
import face_recognition
from datetime import datetime
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, Response
import os
import mimetypes

# ============================
# Flask Setup
# ============================
mimetypes.init(files=[])

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates")
)

# ============================
# Auto Camera Selection
# ============================
def get_camera_source():

    # Laptop Webcam
    cap = cv2.VideoCapture(0)

    if cap.isOpened():
        print("Using laptop webcam")
        return cap

    cap.release()

    # CCTV / IP Camera
    ip_cap = cv2.VideoCapture("rtsp://192.168.1.50:554/stream")

    if ip_cap.isOpened():
        print("Using IP CCTV camera")
        return ip_cap

    ip_cap.release()

    raise Exception("No camera source available")


camera_cap = get_camera_source()

# ============================
# Load Known Faces
# ============================
DATASET_PATH = os.path.join(BASE_DIR, "DataSet")

known_face_encodings = []
known_face_names = []
known_face_rollnos = []

if os.path.exists(DATASET_PATH):

    for student_name in os.listdir(DATASET_PATH):

        student_folder = os.path.join(DATASET_PATH, student_name)

        if os.path.isdir(student_folder):

            for rollno in os.listdir(student_folder):

                roll_folder = os.path.join(student_folder, rollno)

                if os.path.isdir(roll_folder):

                    for img in os.listdir(roll_folder):

                        path = os.path.join(roll_folder, img)

                        if img.lower().endswith((".jpg", ".jpeg", ".png")):

                            try:
                                image = face_recognition.load_image_file(path)

                                small_image = cv2.resize(
                                    image,
                                    (0, 0),
                                    fx=0.25,
                                    fy=0.25
                                )

                                encodings = face_recognition.face_encodings(
                                    small_image,
                                    num_jitters=1
                                )

                                if len(encodings) == 0:
                                    print(f"No face found in {path}")
                                    continue

                                encoding = encodings[0]

                                known_face_encodings.append(encoding)
                                known_face_names.append(student_name)
                                known_face_rollnos.append(rollno)

                                print(f"Loaded: {student_name} ({rollno})")

                            except Exception as e:
                                print(f"Error loading {path}: {e}")

else:
    print("DataSet folder not found")

# ============================
# Master Roll List
# ============================
all_students = [str(i) for i in range(1, 71)]

# ============================
# Database Setup
# ============================
db_path = os.path.join(BASE_DIR, "attendance.db")

conn = sqlite3.connect(db_path, check_same_thread=False)
cursor = conn.cursor()

# Delete old table
cursor.execute("DROP TABLE IF EXISTS attendance")

# Create fresh table
cursor.execute("""
CREATE TABLE attendance (
    name TEXT,
    rollno TEXT,
    date TEXT,
    time TEXT
)
""")

conn.commit()
# ============================
# Excel File Setup
# ============================
excel_file = os.path.join(BASE_DIR, "attendance.xlsx")

if not os.path.exists(excel_file):

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Present Sheet
    ws_present = wb.create_sheet("Present")
    ws_present.append(["Name", "Roll No", "Date", "Time", "Status"])

    # Absent Sheet
    ws_absent = wb.create_sheet("Absent")
    ws_absent.append(["Roll No"])
    ws_absent.append(["name"])

    wb.save(excel_file)

# ============================
# Attendance System
# ============================
attendance_marked = set()

def mark_attendance(name, rollno):

    global attendance_marked

    key = f"{name}_{rollno}"

    if key in attendance_marked:
        return

    now = datetime.now()

    date_string = now.strftime("%d-%m-%Y")
    time_string = now.strftime("%I:%M:%S %p")

    # Save to Database
    cursor.execute(
        "INSERT INTO attendance VALUES (?, ?, ?, ?)",
        (name, rollno, date_string, time_string)
    )

    conn.commit()

    # Save to Excel
    wb = load_workbook(excel_file)

    ws_present = wb["Present"]

    ws_present.append([
        name,
        rollno,
        date_string,
        time_string,
        "Present"
    ])

    wb.save(excel_file)

    attendance_marked.add(key)

    print(f"Attendance marked: {name} ({rollno})")


# ============================
# Save Absentees
# ============================
def save_absentees():

    wb = load_workbook(excel_file)

    if "Absent" not in wb.sheetnames:

        ws_absent = wb.create_sheet("Absent")
        ws_absent.append(["Roll No"])
        ws_absent.append(["Name"])

    else:

        ws_absent = wb["Absent"]

        # Clear previous rows
        if ws_absent.max_row > 1:
            ws_absent.delete_rows(2, ws_absent.max_row)

    # Present students
    cursor.execute("SELECT rollno FROM attendance")

    present_rolls = [row[0] for row in cursor.fetchall()]

    # Absent students
    absent_rolls = [
        roll for roll in all_students
        if roll not in present_rolls
    ]
    present_names = [row[0] for row in cursor.fetchall()]

    absent_names = [
        name for name in all_students
        if name not in present_names
    ]
    # Save absent list
    for roll in absent_rolls:
        ws_absent.append([roll])
    for name in present_names:
        ws_absent.append([name])
    wb.save(excel_file)

    print("Absent roll numbers saved")


# ============================
# Camera Streaming
# ============================
def gen_frames():

    process_frame = True

    while True:

        success, frame = camera_cap.read()

        if not success:
            break

        small_frame = cv2.resize(
            frame,
            (0, 0),
            fx=0.25,
            fy=0.25
        )

        rgb_small_frame = cv2.cvtColor(
            small_frame,
            cv2.COLOR_BGR2RGB
        )

        if process_frame:

            locations = face_recognition.face_locations(
                rgb_small_frame,
                model="hog"
            )

            encodings = face_recognition.face_encodings(
                rgb_small_frame,
                locations
            )

            for encoding, (top, right, bottom, left) in zip(encodings, locations):

                matches = face_recognition.compare_faces(
                    known_face_encodings,
                    encoding,
                    tolerance=0.5
                )

                face_distances = face_recognition.face_distance(
                    known_face_encodings,
                    encoding
                )

                name = "Unknown"
                rollno = "N/A"

                if len(face_distances) > 0:

                    best_match_index = np.argmin(face_distances)

                    if matches[best_match_index]:

                        name = known_face_names[best_match_index]
                        rollno = known_face_rollnos[best_match_index]

                        mark_attendance(name, rollno)

                # Resize coordinates
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4

                # Rectangle
                cv2.rectangle(
                    frame,
                    (left, top),
                    (right, bottom),
                    (0, 255, 0),
                    2
                )

                # Name Text
                cv2.putText(
                    frame,
                    f"{name} ({rollno})",
                    (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.8,
                    (0, 255, 0),
                    2
                )

        process_frame = not process_frame

        ret, buffer = cv2.imencode(".jpg", frame)

        frame = buffer.tobytes()

        yield (
            b'--frame\r\n'
            b'Content-Type: image/jpeg\r\n\r\n' +
            frame +
            b'\r\n'
        )

# ============================
# Flask Routes
# ============================
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/video')
def video():
    return Response(
        gen_frames(),
        mimetype='multipart/x-mixed-replace; boundary=frame'
    )


@app.route('/admin')
def admin_page():

    cursor.execute("SELECT * FROM attendance")

    records = cursor.fetchall()

    save_absentees()

    return render_template(
        'admin.html',
        records=records
    )

@app.route('/absent')
def absent_page():

    # Get present roll numbers
    cursor.execute("SELECT rollno FROM attendance")

    present_rolls = [row[0] for row in cursor.fetchall()]

    absent_students = []

    # Read DataSet folder
    for student_name in os.listdir(DATASET_PATH):

        student_folder = os.path.join(DATASET_PATH, student_name)

        if os.path.isdir(student_folder):

            for rollno in os.listdir(student_folder):

                roll_folder = os.path.join(student_folder, rollno)

                if os.path.isdir(roll_folder):

                    # If student absent
                    if rollno not in present_rolls:

                        absent_students.append({
                            "name": student_name,
                            "rollno": rollno
                        })

    return render_template(
        'absent.html',
        absent_students=absent_students
    )
# ============================
# Run Flask App
# ============================
if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True,
        use_reloader=False
    )